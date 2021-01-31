from pprint import pprint
from openpyxl import load_workbook, Workbook
from django.core.management import BaseCommand
from django.contrib.contenttypes.models import ContentType

from ...models import ApplyProfile, Admission

from abroadin.apps.data.account.models import University, Country, Major
from abroadin.apps.data.applydata.models import Education, RegularLanguageCertificate, Grade, LanguageCertificate, \
    GradeChoices, Publication
from abroadin.apps.data.applydata.utils import convert_pbt_overall_to_ibt_overall, \
    get_ielts_fake_sub_scores_based_on_overall, \
    get_toefl_fake_sub_scores_based_on_overall

LCType = LanguageCertificate.LanguageCertificateType

APPLYPROFILE_CT = ContentType.objects.get(app_label='applyprofile', model='applyprofile')

phd_grade = Grade.objects.filter(name__iexact='ph.d').first()
bachelor_grade = Grade.objects.filter(name__iexact='bachelor').first()
master_grade = Grade.objects.filter(name__iexact='master').first()

default_enroll_year = 2018


def _get_suitable_grade(grade_text):
    if grade_text.lower() == 'doctorate':
        return phd_grade
    elif grade_text.lower() == 'masters':
        return master_grade
    elif grade_text.lower() == 'bachelors':
        return bachelor_grade
    raise Exception("grade is not in right format")


def _get_suitable_lc_type(lc_type_text):
    if lc_type_text.lower() == 'ielts':
        return LCType.IELTS_GENERAL
    elif lc_type_text.lower() == 'ibt' or lc_type_text.lower() == 'pbt':
        return LCType.TOEFL
    raise Exception("Language certificate type is not in right format")


def _get_lc_sub_scores(lc_type, lc_overall):
    if lc_type == LCType.TOEFL:
        return get_toefl_fake_sub_scores_based_on_overall(lc_overall)
    elif lc_type == LCType.IELTS_GENERAL or lc_type == LCType.IELTS_ACADEMIC:
        return get_ielts_fake_sub_scores_based_on_overall(lc_overall)

    raise Exception('wrong lc type')


def detect_outlier(admission_grade, major, destination, scholarship, enroll_year,
                   master_gpa, master_university, bachelor_gpa, bachelor_university, publication_count,
                   lc_type, speaking, listening, writing, reading, overall):
    if publication_count and publication_count > 8:
        raise Exception(
            'outlier data! publication count is enormous. really {} publications?'.format(publication_count))


def insert_into_db(row_id, admission_grade, major, destination, scholarship, enroll_year,
                   master_gpa, master_university, bachelor_gpa, bachelor_university, publication_count,
                   lc_type, speaking, listening, writing, reading, overall
                   ) -> int:
    apply_profile = None
    admission = None
    master_education = None
    bachelor_education = None
    lc = None

    try:
        apply_profile = ApplyProfile.objects.create(
            name='Student #{}'.format(row_id),
            gap=0,
        )

        admission = Admission.objects.create(
            apply_profile=apply_profile,
            grade=admission_grade,
            major=major,
            destination=destination,
            accepted=True,
            scholarship=scholarship,
            enroll_year=enroll_year,
        )

        if master_gpa:
            if master_university is None:
                # TODO handle empty university
                master_university = ''
                pass
            master_education = Education.objects.create(
                grade=GradeChoices.MASTER,
                gpa=master_gpa,
                university=master_university,
                major=major,
                graduate_in=enroll_year,
                content_type=APPLYPROFILE_CT,
                object_id=apply_profile.id,
            )

        if bachelor_gpa:
            if bachelor_university is None:
                # TODO handle empty university
                bachelor_university = ''
                pass
            bachelor_education = Education.objects.create(
                grade=GradeChoices.BACHELOR,
                gpa=bachelor_gpa,
                university=bachelor_university,
                major=major,
                graduate_in=(enroll_year - 4) if master_gpa else enroll_year,
                content_type=APPLYPROFILE_CT,
                object_id=apply_profile.id,
            )

        publications = []
        for i in range(0, publication_count):
            publications.append(
                Publication.objects.create(
                    title=f'Publication #{i}',
                    publish_year=enroll_year - 1,
                    content_type=APPLYPROFILE_CT,
                    object_id=apply_profile.id
                )
            )

        if lc_type:
            pass
            lc = RegularLanguageCertificate.objects.create(
                certificate_type=lc_type,
                speaking=lc,
                content_type=APPLYPROFILE_CT,
                object_id=apply_profile.id,
            )

        return apply_profile.id

    except Exception as e:
        if admission:
            admission.delete()
        if master_education:
            master_education.delete()
        if bachelor_education:
            bachelor_education.delete()
        if publications:
            for publication in publications:
                publication.delete()
        if lc:
            lc.delete()
        if apply_profile:
            apply_profile.delete()

        raise Exception("row {} failed to create. detail: {}".format(row_id, str(e)))


class Command(BaseCommand):
    # https://docs.djangoproject.com/en/3.1/howto/custom-management-commands/
    # https://docs.python.org/3/library/argparse.html#action

    def add_arguments(self, parser):
        parser.add_argument('path', nargs='?', default='')

    def handle(self, *args, **options):
        # https://openpyxl.readthedocs.io/en/stable/usage.html

        input_file_path = options['path']
        wb = load_workbook(input_file_path)

        faulty_rows = []
        warny_rows = []
        succeed_rows = []
        failed_rows = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for i in range(2, sheet.max_row):
                proceed_possible = True

                grade = None
                major = None
                destination = None
                scholarship = None
                master_university = None
                bachelor_university = None
                lc_overall = None
                speaking = None
                listening = None
                writing = None
                reading = None

                ######################
                # row_id
                ######################
                try:
                    row_id = int(sheet.cell(i, 16).value)
                except (ValueError, TypeError):
                    faulty_rows.append((0, 'can not get row id for a row that have {} value in id 16th column'.format(
                        sheet.cell(i, 16).value
                    )))
                    row_id = 0
                    proceed_possible = False

                ######################
                # major
                ######################
                if isinstance(sheet.cell(i, 1).value, int):
                    try:
                        major = Major.objects.get(pk=sheet.cell(i, 1).value)
                    except (Major.DoesNotExist, ValueError, TypeError):
                        # print(i, sheet.cell(i, 1).value)
                        faulty_rows.append((row_id, 'major does not exist'))
                        proceed_possible = False
                else:
                    # print(i, sheet.cell(i, 1).value)
                    faulty_rows.append((row_id, 'major wrong format'))
                    proceed_possible = False

                ######################
                # destination
                ######################
                if isinstance(sheet.cell(i, 2).value, int):
                    try:
                        destination = University.objects.get(pk=sheet.cell(i, 2).value)
                    except (University.DoesNotExist, ValueError, TypeError):
                        # print(i, sheet.cell(i, 2).value)
                        faulty_rows.append((row_id, 'destination does not exist'))
                        proceed_possible = False
                else:
                    # print(i, sheet.cell(i, 2).value)
                    faulty_rows.append((row_id, 'destination wrong format'))
                    proceed_possible = False

                ######################
                # grade
                ######################
                if isinstance(sheet.cell(i, 3).value, str):
                    try:
                        grade = _get_suitable_grade(sheet.cell(i, 3).value)
                    except Exception as e:
                        # print(i, sheet.cell(i, 3).value)
                        faulty_rows.append((row_id, str(e)))
                        proceed_possible = False
                else:
                    # print(i, sheet.cell(i, 3).value)
                    faulty_rows.append((row_id, 'grade wrong format'))
                    proceed_possible = False

                ######################
                # scholarship
                ######################
                if sheet.cell(i, 4).value is None:
                    scholarship = 0
                    # print(i, sheet.cell(i, 4).value)
                else:
                    try:
                        scholarship = int(sheet.cell(i, 4).value)
                    except (TypeError, ValueError):
                        # print(i, sheet.cell(i, 4).value)
                        faulty_rows.append((row_id, 'fund wrong format'))
                        proceed_possible = False

                ######################
                # bachelor gpa
                ######################
                if sheet.cell(i, 5).value is None:
                    bachelor_gpa = None
                    # print(i, sheet.cell(i, 5).value)
                else:
                    try:
                        bachelor_gpa = float(sheet.cell(i, 5).value)
                    except (TypeError, ValueError):
                        # print(i, sheet.cell(i, 5).value)
                        bachelor_gpa = None
                        faulty_rows.append((row_id, 'bachelor_gpa wrong format'))
                        proceed_possible = False

                ######################
                # bachelor university
                ######################
                if bachelor_gpa:
                    if isinstance(sheet.cell(i, 6).value, int):
                        try:
                            bachelor_university = University.objects.get(pk=sheet.cell(i, 6).value)
                        except (University.DoesNotExist, ValueError, TypeError):
                            # print(i, sheet.cell(i, 6).value)
                            bachelor_university = None
                            faulty_rows.append((row_id, 'bachelor_gpa has value but bachelor_university does not exist'))
                            proceed_possible = False
                    else:
                        # print(i, sheet.cell(i, 6).value)
                        bachelor_university = None
                        faulty_rows.append((row_id, 'bachelor_gpa has value but bachelor_university is string or empty'))
                        proceed_possible = False

                ######################
                # master gpa
                ######################
                if sheet.cell(i, 7).value is None:
                    master_gpa = None
                    # print(i, sheet.cell(i, 7).value)
                else:
                    try:
                        master_gpa = float(sheet.cell(i, 7).value)
                    except (TypeError, ValueError):
                        # print(i, sheet.cell(i, 7).value)
                        master_gpa = None
                        faulty_rows.append((row_id, 'master_gpa wrong format'))
                        proceed_possible = False

                ######################
                # master university
                ######################
                if master_gpa:
                    if isinstance(sheet.cell(i, 8).value, int):
                        try:
                            master_university = University.objects.get(pk=sheet.cell(i, 8).value)
                        except (University.DoesNotExist, ValueError, TypeError):
                            # print(i, sheet.cell(i, 8).value)
                            master_university = None
                            faulty_rows.append((row_id, 'master_gpa has value but master_university does not exist'))
                            proceed_possible = False
                    else:
                        # print(i, sheet.cell(i, 8).value)
                        master_university = None
                        faulty_rows.append((row_id, 'master_gpa has value but master_university is string or empty'))
                        proceed_possible = False

                ######################
                # LC Type
                ######################
                if isinstance(sheet.cell(i, 9).value, str):
                    try:
                        lc_type_text = sheet.cell(i, 9).value
                        lc_type = _get_suitable_lc_type(lc_type_text)
                    except Exception as e:
                        lc_type = None
                        lc_type_text = None
                        faulty_rows.append((row_id, str(e)))
                        proceed_possible = False
                        # print(i, sheet.cell(i, 9).value)
                else:
                    lc_type = None
                    lc_type_text = None
                    # print(i, sheet.cell(i, 9).value)

                ######################
                # LC overall
                ######################
                if lc_type is not None:
                    try:
                        lc_overall = float(sheet.cell(i, 10).value)
                        if lc_type_text.lower() == 'pbt':
                            try:
                                lc_overall = convert_pbt_overall_to_ibt_overall(lc_overall)
                                lc_type_text = 'ibt'
                            except Exception as e:
                                faulty_rows.append((row_id, str(e)))
                                proceed_possible = False
                    except (ValueError, TypeError):
                        print(sheet.cell(i, 10).value)
                        lc_overall = None
                        faulty_rows.append((row_id, 'language certificate overall score is in wrong format'))
                        proceed_possible = False

                if lc_type and lc_overall:
                    writing, reading, speaking, listening = _get_lc_sub_scores(lc_type, lc_overall)
                    # try:
                    #     writing, reading, speaking, listening = _get_lc_sub_scores(lc_type, lc_overall)
                    # except Exception as e:
                    #     faulty_rows.append((row_id, 'language certificate overall score is in wrong format, can not'
                    #                                 ' generate fake scores. detail: {}'.format(str(e))))
                    #     proceed_possible = False

                # # TODO Determine certificate score :/
                # ######################
                # # LC
                # ######################
                # if lc_type is not None:
                #     try:
                #         lc_q = float(sheet.cell(i, 11).value)
                #     except (ValueError, TypeError):
                #         lc_q = None
                #         faulty_rows.append((row_id, 'language certificate overall score is in wrong format'))
                #         proceed_possible = False
                #
                # # TODO Determine certificate score :/
                # ######################
                # # LC a
                # ######################
                # if lc_type is not None:
                #     try:
                #         lc_a = float(sheet.cell(i, 12).value)
                #     except (ValueError, TypeError):
                #         lc_a = None
                #         faulty_rows.append((row_id, 'language certificate overall score is in wrong format'))
                #         proceed_possible = False
                #
                # # TODO Determine certificate score :/
                # ######################
                # # LC v
                # ######################
                # if lc_type is not None:
                #     try:
                #         lc_v = float(sheet.cell(i, 13).value)
                #     except (ValueError, TypeError):
                #         lc_v = None
                #         faulty_rows.append((row_id, 'language certificate overall score is in wrong format'))
                #         proceed_possible = False

                ######################
                # pub count
                ######################
                if sheet.cell(i, 14).value is None:
                    pub_count = 0
                    warny_rows.append((row_id, 'publication count is empty'))
                    # print(sheet.cell(i, 14).value)
                else:
                    try:
                        pub_count = int(sheet.cell(i, 14).value)
                        # print(sheet.cell(i, 14).value)
                    except (ValueError, TypeError):
                        pub_count = None
                        faulty_rows.append((row_id, 'publication count is in wrong format'))
                        proceed_possible = False
                        # print(sheet.cell(i, 14).value)

                ######################
                # enroll year
                ######################
                try:
                    enroll_year = int(sheet.cell(i, 15).value)
                except (ValueError, TypeError):
                    enroll_year = default_enroll_year
                    faulty_rows.append((row_id, 'enroll_year wrong format'))
                    proceed_possible = False

                try:
                    detect_outlier(grade, major, destination, scholarship, enroll_year,
                                   master_gpa, master_university, bachelor_gpa, bachelor_university, pub_count,
                                   lc_type, speaking, listening, writing, reading, lc_overall)
                except Exception as e:
                    faulty_rows.append((row_id, str(e)))

                # ###################################### Create objects ###############################################

                if proceed_possible:
                    try:

                        db_id = insert_into_db(row_id, grade, major, destination, scholarship, enroll_year,
                                               master_gpa, master_university, bachelor_gpa, bachelor_university, pub_count,
                                               lc_type, speaking, listening, writing, reading, lc_overall)

                        succeed_rows.append((row_id, "inserted successfully into db with object id: {}".format(db_id)))
                    except Exception as e:
                        failed_rows.append((row_id, str(e)))
                    # print(i, str(major).strip(), destination)

                else:
                    failed_rows.append((row_id, "failed to insert. more details in faulty rows"))


        # print('succeed rows')
        # pprint(succeed_rows)
        # print(failed_rows)
        # pprint(failed_rows)
        #
        # print('faulty_rows')
        # pprint(faulty_rows)
        # print('warny_rows')
        # pprint(warny_rows)

        with open('/'.join(input_file_path.split('/')[:-1] + ['succeed rows']), 'w') as f:
            for entry in succeed_rows:
                f.write(str(entry) + '\n')

        with open('/'.join(input_file_path.split('/')[:-1] + ['failed rows']), 'w') as f:
            for entry in failed_rows:
                f.write(str(entry) + '\n')

        with open('/'.join(input_file_path.split('/')[:-1] + ['faulty rows']), 'w') as f:
            for entry in faulty_rows:
                f.write(str(entry) + '\n')

        with open('/'.join(input_file_path.split('/')[:-1] + ['warny rows']), 'w') as f:
            for entry in warny_rows:
                f.write(str(entry) + '\n')

        print("Finished. Check xlsx file directory for output files.")
